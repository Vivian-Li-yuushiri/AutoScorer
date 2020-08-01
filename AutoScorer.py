from tkinter import *
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
from io import StringIO

import threading
import sys
import time
import os

import numpy as np
import sqlite3
import mne

class NullIO(StringIO):
	def write(self, txt):
		pass

#Formats the edf for the algorithm and stores it as a class
class FormatData(object):
	def __init__(self, filePath = ""):
		self.filePath = filePath
		self.fileName = filePath.split("/")[-1][:-4]

	def set_file_path(self, filePath):
		self.filePath = filePath
		self.fileName = filePath.split("/")[-1][:-4]

	def prepare_raws(self):
                #only accepts edfs if the form of two eeg channels and one emg channel
		def load_edf():
			def set_channel_types(raw):
				names = raw.ch_names
				ch_map = {names[0]: "eeg", names[1]: "eeg", names[2]: "emg"}
				raw.set_channel_types(ch_map)
				return raw
			self.raw = mne.io.read_raw_edf(self.filePath, preload = True)
			self.raw = self.raw.pick_types(eeg = True, emg = True)
			self.raw = set_channel_types(self.raw)
                #separates into 10 second epochs for scoring
		def get_epochs():
			self.epochs = []
			length = len(self.raw)
			for i in range(0, length - 10000, 10000):
				data = self.raw.get_data(start = i, stop = i+10000)
				self.epochs.append(data)
			self.epochs.append(self.raw.get_data(start = length - 10000))
                #save the frequencies of epochs to an array
		def get_psds():
			self.psds = []
			length = len(self.epochs)
			for i in range(0, length):
				psd = mne.time_frequency.psd_array_multitaper(self.epochs[i][:2], 1000, fmin = 0, fmax = 8)
				self.psds.append(psd)
                #convert epochs to an array of raw data
		def epochs_to_raws():
			self.raws = []
			for epoch in self.epochs:
				self.raws.append(mne.io.RawArray(epoch, self.raw.info, verbose = None))
		print("Preparing stuff...\n")
		sys.stdout = NullIO()
		load_edf()
		get_epochs()
		get_psds()
		epochs_to_raws()
		sys.stdout = sys.__stdout__
		print("Done!\n")

        #convert the psds and raw data into a readable format through an algorithm using thresholds
	def get_all_data(self):
		def find_edf_thresholds():
			eeg_data = np.abs(self.raw.get_data(picks = ["eeg"]))
			emg_data = np.abs(self.raw.get_data(picks = ["emg"]))
			self.eeg_threshold = eeg_data.mean()*1.5
			self.emg_threshold = emg_data.mean()*3

		def find_psd_thresholds(psd):
			return psd.max()/1.5

		def get_data(epoch, psd, psd_threshold):
			eeg_data = np.abs(epoch[0:2])
			emg_data = np.abs(epoch[2])
			eeg_psd = np.abs(psd)
			eeg_count = [0]
			eeg_count *= len(eeg_data)
			emg_count = 0
			psd_count = {"0-4": [0, 0], "4-8": [0, 0]}

			for eeg_i in range(len(eeg_data)):
				for value in eeg_data[eeg_i]:
					if value > self.eeg_threshold:
						eeg_count[eeg_i] += 1

			for value in emg_data:
				if value > self.emg_threshold:
					emg_count += 1

			for psd_i in range(len(eeg_psd)):
				for value in eeg_psd[psd_i][0:41]:
					if value > psd_threshold:
						psd_count["0-4"][psd_i] += 1
				for value in eeg_psd[psd_i][41:]:
					if value > psd_threshold:
						psd_count["4-8"][psd_i] += 1

			return np.mean(eeg_count), emg_count, np.mean(psd_count["0-4"]), np.mean(psd_count["4-8"])

		self.datas = []
		find_edf_thresholds()

		for i in range(len(self.epochs)):
			if i % (len(self.epochs)//50) == 0:
				print("[" + "█"*(i//(len(self.epochs)//50)) + " "*(50 - (i//(len(self.epochs)//50))) + "] (%d/50)" % (i//(len(self.epochs)//50)))
			new_psd = self.psds[i]
			psd_threshold = find_psd_thresholds(new_psd[0])
			new_data = get_data(self.epochs[i], new_psd[0], psd_threshold)
			self.datas.append(new_data)
		print("Done!\n")

#Algorithm to classify formatted data into different sleep states
class EDFScorer(object):
	def __init__(self, filePath = "", destiPath = ""):
		self.Data = FormatData()
		self.filePath = filePath
		self.destiPath = destiPath

	def set_file_path(self, filePath):
		self.filePath = filePath
		self.Data.set_file_path(filePath)

	def set_desti_path(self, destiPath):
		self.destiPath = destiPath

	def get_raws(self):
		self.Data.prepare_raws()
		self.Data.get_all_data()

        #manual fixes based on rules for sleep scoring
	def fix(self):
		for i in range(len(self.evals)-2):
			if self.evals[i] == 1 and self.evals[i+1] == 3:
				self.evals[i+1] = 1
			if self.evals[i] == 2 and self.evals[i+1] == 3 and self.evals[i+2] == 1:
				self.evals[i+1] = 1
			if self.evals[i] == 1 and self.evals[i+1] == 2 and self.evals[i+2] == 1:
				self.evals[i+1] = 1
			if self.evals[i] == 2 and self.evals[i+1] == 3 and self.evals[i+2] == 2:
				self.evals[i+1] = 2
			if self.evals[i] == 3 and self.evals[i+1] == 2 and self.evals[i+2] == 3:
				self.evals[i+1] = 3

	def get_eval(self):
		def find_data_threshold():
			data = np.array(self.Data.datas)
			eeg_data = []
			emg_data = data[:, 1]
			for epoch in self.Data.datas:
				eeg_data.append(np.mean(epoch[0]))
			self.eeg_threshold = np.mean(eeg_data)
			self.emg_threshold = np.mean(emg_data)
		def first_eval(data):
			eeg_data = data[0]
			emg_data = data[1]
			psd_nrem = data[2]
			psd_rem = data[3]

			if emg_data > self.emg_threshold:
				return 1 #wake
			else:
				if eeg_data > self.eeg_threshold:
					return 2 #nrem
				else:
					if psd_rem > psd_nrem:
						if emg_data < self.emg_threshold/1.5:
							if eeg_data > self.eeg_threshold:
								return 2
							return 3 #rem
						return 1
					else:
						if eeg_data > self.eeg_threshold/1.1:
							if emg_data > self.emg_threshold/1.5:
								return 1
							return 2
						else:
							if emg_data < self.emg_threshold/1.5:
								return 3
							return 1
		self.evals = []
		find_data_threshold()
		for new_data in self.Data.datas:
			new_eval = first_eval(new_data)
			self.evals.append(new_eval)
		self.fix()

        #saves results as an excel file
	def to_excel(self):
		if os.path.exists(self.destiPath + "results.xlsx"):
			wb = load_workbook(self.destiPath + "results.xlsx")
			ws = wb.create_sheet(title = self.Data.fileName)
			for item in self.evals:
				ws.append((item,))
		else:
			wb = Workbook()
			ws = wb.active
			ws.title = self.Data.fileName
			for item in self.evals:
				ws.append((item,))
		wb.save(self.destiPath + "results.xlsx")

        #compares results with a pre-scored sql file (for Broad)
	def compare_data(self):
		def evals_from_sql():
			conn = sqlite3.connect(self.Data.filePath[:-4] + ".db3")
			cur = conn.cursor()
			cur.execute("SELECT start_time_seconds, score FROM sleep_scores_table ORDER BY start_time_seconds ASC")
			rows = cur.fetchall()

			evals = []
			for row in rows:
				evals.append(row[1])
			return evals
		data_1 = self.evals
		data_2 = evals_from_sql()
		if len(data_1) != len(data_2):
			print("Error!\n")
			return

		sim = 0
		for i in range(len(data_1)):
			if data_1[i] == data_2[i]:
				sim += 1
		print(round(sim/len(data_1), 4) * 100)

#creates the graphics interface to run everything
class GUI(object):
	def redirector(self, inputStr):
		self.results.delete("end linestart", "end")
		self.results.insert("end", inputStr)

	def __init__(self):
		self.root = Tk()
		self.root.title("Auto-sleep Scoring")
		self.root.resizable(width = False, height = False)

		self.EDFScorer = EDFScorer()

		self.filePath = StringVar()
		self.destiPath = StringVar()
		self.pathLabel = Label(self.root, textvariable = self.filePath)
		self.destiLabel = Label(self.root, textvariable = self.destiPath)
		self.browseEdf = Button(self.root, text = "Browse EDF", command = self.edfBrowse)
		self.browseEnd = Button(self.root, text = "Final Location", command = self.destiBrowse)
		self.runButton = Button(self.root, text = "Run", command = lambda: self.run(self.filePath.get()))
		self.resultsFrame = Frame(self.root)
		self.scrollBar = Scrollbar(self.resultsFrame)
		self.results = Text(self.resultsFrame, yscrollcommand = self.scrollBar.set)

	def pack(self):
		self.pathLabel.pack()
		self.destiLabel.pack()
		self.browseEdf.pack()
		self.browseEnd.pack()
		self.runButton.pack()
		self.resultsFrame.pack()
		self.scrollBar.pack(side = RIGHT, fill = Y)
		self.results.pack()
		self.scrollBar.config(command = self.results.yview)
		self.root.mainloop()

        
	def edfBrowse(self):
	    filename = filedialog.askopenfilename(filetypes = (("edf files", "*.edf"), ("all files", "*.*")))
	    self.filePath.set(filename)
	    self.EDFScorer.set_file_path(filename)

	def destiBrowse(self):
		directory = filedialog.askdirectory() + "/"
		self.destiPath.set(directory)
		self.EDFScorer.set_desti_path(directory)

	def save_eval(self):
		start = time.time()
		self.EDFScorer.get_raws()
		self.EDFScorer.get_eval()
		self.EDFScorer.to_excel()
		end = time.time()

		print("Time taken: " + str(round(end - start, 2)) + " seconds")

	def run(self, file):
		sys.stdout.write = self.redirector
		t = threading.Thread(target = lambda: self.save_eval())
		t.deamon = True
		t.start()

gui = GUI()
gui.pack()
